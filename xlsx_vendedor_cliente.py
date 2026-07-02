"""Gerador de Excel - Relatorio Vendedor-Cliente OTHIL.

Arquitetura:
  historico (JSON, uma vez por mes):
    ant_ano / ant_mes por vendedor: clientes dict + TOTAL dict

  atual (toda sexta):
    clientes_atual  = parse_e_agregar(pdf_vendedor_cliente)
    totais_atual    = parse_totais_vendedor(pdf_lucratividade)['vendedores']

Regras:
  MC_RS = Fat - Custo_PDF
  MC%   = MC_RS / Custo * 100
  Resultado Real = MC% + 15pp   <- coluna MARGEM %
  Luca EXCLUIDO.
"""
from __future__ import annotations

import io
import json
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from parsers_diario import parse_relatorio_diario

_EXCLUIDOS = {'Luca'}

VENDOR_TAB = {
    'Afanais':   'AFANAIS',
    'Claudia':   'CLAUDIA',
    'Dora':      'DORA',
    'Farley':    'FARLEY',
    'Juliana':   'JULIANA',
    'Luciano':   'LUCIANO',
    'Reginaldo': 'REGINALDO',
    'Roni':      'RONISTONIS',
}
VENDOR_TITLE = {**VENDOR_TAB, 'Roni': 'RONI'}
VENDOR_ORDER = ['Afanais', 'Dora', 'Farley', 'Luciano',
                'Reginaldo', 'Roni', 'Claudia', 'Juliana']

def _fill(h): return PatternFill('solid', fgColor=h)
def _font(bold=False, color='000000', size=9):
    return Font(bold=bold, color=color, size=size, name='Calibri')

AL_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL_L = Alignment(horizontal='left',   vertical='center')
AL_R = Alignment(horizontal='right',  vertical='center')

FILL_TIT     = _fill('1F4E79')
FONT_TIT     = _font(bold=True, color='FFFFFF', size=11)
FILL_CLI_HDR = _fill('4472C4')
FONT_CLI_HDR = _font(bold=True, color='FFFFFF')
FILL_ROW_A   = _fill('FFFFFF')
FILL_ROW_B   = _fill('F5F5F5')
FILL_TOT     = _fill('BDD7EE')
FONT_DATA    = _font(size=9)
FONT_NEG     = _font(color='C00000', size=9)
FONT_TOT     = _font(bold=True, size=9)
FONT_TOT_N   = _font(bold=True, color='C00000', size=9)

_P_FILL_HDR = {
    'ant_ano': _fill('2F75B6'), 'ant_mes': _fill('5B9BD5'),
    'meta':    _fill('375623'), 'atual':   _fill('1F4E79'),
}
_P_FONT_HDR = {k: _font(bold=True, color='FFFFFF')
               for k in ('ant_ano', 'ant_mes', 'meta', 'atual')}
_P_FILL_DAT = {
    'ant_ano': _fill('DEEAF1'), 'ant_mes': _fill('EBF3FB'),
    'meta':    _fill('E2EFDA'), 'atual':   _fill('FFF2CC'),
}

_P_COL      = {'ant_ano': 2, 'ant_mes': 5, 'meta': 8, 'atual': 11}
_GERAL_PERS = ['ant_ano', 'ant_mes', 'atual']
_GERAL_COL  = {'ant_ano': 2, 'ant_mes': 5, 'atual': 8}


def _calc(fat, custo):
    mc_rs  = fat - custo
    mc_pct = mc_rs / custo * 100 if custo else 0.0
    return round(mc_rs, 2), round(mc_pct, 2), round(mc_pct + 15, 2)


def _periodo_labels(ref_date):
    M = ['jan','fev','mar','abr','mai','jun','jul','ago','set','out','nov','dez']
    m, y = ref_date.month, ref_date.year
    m2 = m - 1 if m > 1 else 12
    y2 = y if m > 1 else y - 1
    return {
        'ant_ano': f"{M[m-1]}./{y-1}",
        'ant_mes': f"{M[m2-1]}./{y2}",
        'meta':    'META',
        'atual':   f"ATUAL ({M[m-1].upper()}/{y})",
        '_ref':    f"{M[m-1].upper()}/{y}",
    }


def _st(c, fill=None, font=None, align=None, fmt=None):
    if fill:  c.fill          = fill
    if font:  c.font          = font
    if align: c.alignment     = align
    if fmt:   c.number_format = fmt


def _dash(ws, r, col, fill, p_key=None):
    f = _P_FILL_DAT[p_key] if p_key else fill
    for off in range(3):
        c = ws.cell(r, col + off, 'x')
        c.value = '—'
        _st(c, fill=f, font=FONT_DATA, align=AL_C)


def parse_e_agregar(file_objs):
    """Parse um ou mais PDFs Vendedor-Cliente e agrega tudo.

    Aceita um unico file-like object OU uma lista deles (multi-upload).
    Retorna {vendedor: {cliente: {vol,fat,custo,mc_rs,mc_pct,resultado_real}}}
    """
    if file_objs is None:
        return {}
    if not isinstance(file_objs, (list, tuple)):
        file_objs = [file_objs]

    raw = defaultdict(lambda: defaultdict(lambda: {'vol': 0.0, 'fat': 0.0, 'custo': 0.0}))

    for file_obj in file_objs:
        if file_obj is None:
            continue
        try:
            result = parse_relatorio_diario(file_obj)
            itens  = result['itens']
        except Exception:
            continue
        for it in itens:
            if it.get('vendedor') in _EXCLUIDOS:
                continue
            v, c = it['vendedor'], it['cliente_nome']
            raw[v][c]['vol']   += it.get('qtd', 0)
            raw[v][c]['fat']   += it.get('faturamento', 0)
            raw[v][c]['custo'] += it.get('custo_total', 0)

    out = {}
    for v, clientes in raw.items():
        out[v] = {}
        for c, d in clientes.items():
            mc_rs, mc_pct, res = _calc(d['fat'], d['custo'])
            out[v][c] = {
                'vol': round(d['vol']), 'fat': round(d['fat'], 2),
                'custo': round(d['custo'], 2), 'mc_rs': mc_rs,
                'mc_pct': mc_pct, 'resultado_real': res,
            }
    return out


def agregar_totais_historicos(file_obj):
    """Soma todos os itens por vendedor -> total real (inclui todos os clientes)."""
    if file_obj is None:
        return {}
    try:
        result = parse_relatorio_diario(file_obj)
        itens  = result['itens']
    except Exception:
        return {}

    raw = defaultdict(lambda: {'vol': 0.0, 'fat': 0.0, 'custo': 0.0})
    for it in itens:
        if it.get('vendedor') in _EXCLUIDOS:
            continue
        v = it['vendedor']
        raw[v]['vol']   += it.get('qtd', 0)
        raw[v]['fat']   += it.get('faturamento', 0)
        raw[v]['custo'] += it.get('custo_total', 0)

    out = {}
    for v, d in raw.items():
        mc_rs, mc_pct, res = _calc(d['fat'], d['custo'])
        out[v] = {
            'vol': round(d['vol']), 'fat': round(d['fat'], 2),
            'custo': round(d['custo'], 2), 'mc_rs': mc_rs,
            'mc_pct': mc_pct, 'resultado_real': res,
        }
    return out


def salvar_historico(pdf_ant_ano, pdf_ant_mes, ref_date):
    """Parsa os 2 PDFs historicos e retorna bytes do JSON."""
    labels = _periodo_labels(ref_date)

    clientes_ant_ano = parse_e_agregar(pdf_ant_ano)
    clientes_ant_mes = parse_e_agregar(pdf_ant_mes)
    totais_ant_ano   = agregar_totais_historicos(pdf_ant_ano)
    totais_ant_mes   = agregar_totais_historicos(pdf_ant_mes)

    def _build(clientes, totais):
        out = {}
        for v in set(clientes) | set(totais):
            if v in _EXCLUIDOS:
                continue
            out[v] = {'clientes': clientes.get(v, {}), 'TOTAL': totais.get(v, {})}
        return out

    historico = {
        'ref_label': labels['_ref'],
        'gerado_em': datetime.now().isoformat(timespec='seconds'),
        'ant_ano': {
            'label':     labels['ant_ano'],
            'vendedores': _build(clientes_ant_ano, totais_ant_ano),
        },
        'ant_mes': {
            'label':     labels['ant_mes'],
            'vendedores': _build(clientes_ant_mes, totais_ant_mes),
        },
    }
    return json.dumps(historico, ensure_ascii=False, indent=2).encode('utf-8')


def carregar_historico(json_bytes):
    return json.loads(json_bytes.decode('utf-8'))


def _build_headers(ws, ref_label, labels, n_cols, p_col_map, title_name):
    last = get_column_letter(n_cols)
    ws.merge_cells(f'A1:{last}1')
    c = ws.cell(1, 1, f'RELATORIO VENDEDOR-CLIENTE -- {title_name}  |  Ref: {ref_label}')
    _st(c, fill=FILL_TIT, font=FONT_TIT, align=AL_C)
    ws.row_dimensions[1].height = 22

    ws.cell(2, 1, 'CLIENTE')
    _st(ws.cell(2, 1), fill=FILL_CLI_HDR, font=FONT_CLI_HDR, align=AL_C)
    for p, c0 in p_col_map.items():
        ws.merge_cells(start_row=2, start_column=c0, end_row=2, end_column=c0+2)
        c = ws.cell(2, c0, labels[p])
        _st(c, fill=_P_FILL_HDR[p], font=_P_FONT_HDR[p], align=AL_C)
    ws.row_dimensions[2].height = 28

    _st(ws.cell(3, 1), fill=FILL_CLI_HDR, font=FONT_CLI_HDR, align=AL_C)
    for p, c0 in p_col_map.items():
        for off, lbl in enumerate(['VOLUME', 'FATURAMENTO', 'MARGEM %']):
            c = ws.cell(3, c0+off, lbl)
            _st(c, fill=_P_FILL_HDR[p], font=_P_FONT_HDR[p], align=AL_C)
    ws.row_dimensions[3].height = 14


def _write_data(ws, r, c0, d, p_key, bold=False):
    if not d:
        _dash(ws, r, c0, None, p_key)
        return
    res_dec = d['resultado_real'] / 100
    fnt_m = (FONT_TOT_N if res_dec < 0 else FONT_TOT) if bold else (FONT_NEG if res_dec < 0 else FONT_DATA)
    fnt_n = FONT_TOT if bold else FONT_DATA
    fill  = _P_FILL_DAT[p_key]

    cv = ws.cell(r, c0,   d['vol'])
    cf = ws.cell(r, c0+1, d['fat'])
    cm = ws.cell(r, c0+2, res_dec)
    cv.number_format = '#,##0'
    cf.number_format = '#,##0.00'
    cm.number_format = '0.00%'
    for off in range(3):
        _st(ws.cell(r, c0+off), fill=fill, align=AL_R)
    cv.font = fnt_n
    cf.font = fnt_n
    cm.font = fnt_m


def _write_total_row(ws, r, c0, d, p_key):
    if not d:
        _dash(ws, r, c0, FILL_TOT, p_key)
        return
    res_dec = d['resultado_real'] / 100
    fnt_m   = FONT_TOT_N if res_dec < 0 else FONT_TOT

    cv = ws.cell(r, c0,   d['vol'])
    cf = ws.cell(r, c0+1, d['fat'])
    cm = ws.cell(r, c0+2, res_dec)
    cv.number_format = '#,##0'
    cf.number_format = '#,##0.00'
    cm.number_format = '0.00%'
    for off in range(3):
        _st(ws.cell(r, c0+off), fill=FILL_TOT, align=AL_R)
    cv.font = FONT_TOT
    cf.font = FONT_TOT
    cm.font = fnt_m


def _build_vendedor_sheet(ws, title_name, clientes_por_periodo, totais_por_periodo,
                          meta_vend, labels, ref_label):
    ws.freeze_panes = 'B4'
    ws.column_dimensions['A'].width = 34
    for p, c0 in _P_COL.items():
        ws.column_dimensions[get_column_letter(c0)].width   = 9
        ws.column_dimensions[get_column_letter(c0+1)].width = 14
        ws.column_dimensions[get_column_letter(c0+2)].width = 10

    _build_headers(ws, ref_label, labels, 13, _P_COL, title_name)

    all_cli = set()
    for pdata in clientes_por_periodo.values():
        all_cli.update(pdata.keys())
    for ck in meta_vend:
        if not any(c.upper() == ck for c in all_cli):
            all_cli.add(ck)
    all_cli_sorted = sorted(all_cli)

    for i, cli in enumerate(all_cli_sorted):
        r = 4 + i
        ws.row_dimensions[r].height = 14
        fill_r = FILL_ROW_A if i % 2 == 0 else FILL_ROW_B
        c = ws.cell(r, 1, cli)
        _st(c, fill=fill_r, font=FONT_DATA, align=AL_L)

        for p_key, c0 in _P_COL.items():
            if p_key == 'meta':
                m = meta_vend.get(cli.upper())
                if m:
                    cv = ws.cell(r, c0,   m.get('vol'))
                    cf = ws.cell(r, c0+1, m.get('fat'))
                    cm = ws.cell(r, c0+2, m.get('margem', 0.15))
                    cv.number_format = '#,##0'
                    cf.number_format = '#,##0.00'
                    cm.number_format = '0.00%'
                    for off in range(3):
                        _st(ws.cell(r, c0+off),
                            fill=_P_FILL_DAT['meta'], font=FONT_DATA, align=AL_R)
                else:
                    _dash(ws, r, c0, fill_r, 'meta')
            else:
                d = clientes_por_periodo.get(p_key, {}).get(cli)
                _write_data(ws, r, c0, d, p_key)

    tot_row = 4 + len(all_cli_sorted)
    ws.row_dimensions[tot_row].height = 16
    c = ws.cell(tot_row, 1, 'TOTAL')
    _st(c, fill=FILL_TOT, font=FONT_TOT, align=AL_L)

    for p_key, c0 in _P_COL.items():
        if p_key == 'meta':
            t_vol = sum(m.get('vol') or 0 for m in meta_vend.values())
            t_fat = sum(m.get('fat') or 0 for m in meta_vend.values())
            cv = ws.cell(tot_row, c0,   t_vol or None)
            cf = ws.cell(tot_row, c0+1, t_fat or None)
            cm = ws.cell(tot_row, c0+2, 0.15)
            cv.number_format = '#,##0'
            cf.number_format = '#,##0.00'
            cm.number_format = '0.00%'
            for off in range(3):
                _st(ws.cell(tot_row, c0+off), fill=FILL_TOT, font=FONT_TOT, align=AL_R)
        else:
            _write_total_row(ws, tot_row, c0, totais_por_periodo.get(p_key), p_key)


def _build_geral_sheet(ws, geral_data, labels, ref_label):
    ws.freeze_panes = 'B4'
    ws.column_dimensions['A'].width = 18
    for p in _GERAL_PERS:
        c0 = _GERAL_COL[p]
        ws.column_dimensions[get_column_letter(c0)].width   = 9
        ws.column_dimensions[get_column_letter(c0+1)].width = 14
        ws.column_dimensions[get_column_letter(c0+2)].width = 10

    p_col_g = {k: _GERAL_COL[k] for k in _GERAL_PERS}
    _build_headers(ws, ref_label, labels, 10, p_col_g, 'CONSOLIDADO')
    ws.cell(1, 1).value = f'GERAL  |  Ref: {ref_label}'

    vend_keys = [v for v in VENDOR_ORDER if v in geral_data]
    for v in sorted(geral_data):
        if v not in vend_keys:
            vend_keys.append(v)

    for i, vend in enumerate(vend_keys):
        r = 4 + i
        fill_r = FILL_ROW_A if i % 2 == 0 else FILL_ROW_B
        ws.row_dimensions[r].height = 14
        c = ws.cell(r, 1, VENDOR_TAB.get(vend, vend.upper()))
        _st(c, fill=fill_r, font=FONT_DATA, align=AL_L)
        for p in _GERAL_PERS:
            _write_data(ws, r, _GERAL_COL[p], geral_data[vend].get(p), p)

    tot_row = 4 + len(vend_keys)
    ws.row_dimensions[tot_row].height = 16
    c = ws.cell(tot_row, 1, 'TOTAL')
    _st(c, fill=FILL_TOT, font=FONT_TOT, align=AL_L)

    for p in _GERAL_PERS:
        c0   = _GERAL_COL[p]
        rows = [geral_data[v].get(p) for v in vend_keys if geral_data[v].get(p)]
        if rows:
            t_vol   = sum(d['vol']   for d in rows)
            t_fat   = sum(d['fat']   for d in rows)
            t_custo = sum(d['custo'] for d in rows)
            _, _, res = _calc(t_fat, t_custo)
            grand = {'vol': t_vol, 'fat': t_fat, 'custo': t_custo, 'resultado_real': res}
            _write_total_row(ws, tot_row, c0, grand, p)
        else:
            _dash(ws, tot_row, c0, FILL_TOT, p)


def gerar_xlsx(historico, pdf_clientes_atual, totais_atual,
               meta_xlsx_bytes=None, ref_date=None):
    """Gera o Excel Vendedor-Cliente e retorna bytes.

    historico          : dict de carregar_historico()
    pdf_clientes_atual : file-like (PDF Vendedor-Cliente atual)
    totais_atual       : dict de parse_totais_vendedor()['vendedores']
    meta_xlsx_bytes    : bytes do xlsx anterior (opcional, para META)
    ref_date           : datetime de referencia (default: hoje)
    """
    if ref_date is None:
        ref_date = datetime.today()

    labels    = _periodo_labels(ref_date)
    ref_label = labels.pop('_ref')

    ant_ano_h = historico.get('ant_ano', {})
    ant_mes_h = historico.get('ant_mes', {})
    if ant_ano_h.get('label'): labels['ant_ano'] = ant_ano_h['label']
    if ant_mes_h.get('label'): labels['ant_mes'] = ant_mes_h['label']

    clientes_atual = parse_e_agregar(pdf_clientes_atual)

    meta_global = {}
    if meta_xlsx_bytes:
        try:
            wb_m = openpyxl.load_workbook(io.BytesIO(meta_xlsx_bytes), data_only=True)
            for ws_m in wb_m.worksheets:
                if ws_m.title.strip().upper() == 'GERAL':
                    continue
                d = {}
                for row in ws_m.iter_rows(min_row=4, values_only=True):
                    if not row or not row[0]:
                        continue
                    cli = str(row[0]).strip()
                    if cli.upper() == 'TOTAL':
                        continue
                    vol = row[7] if len(row) > 7 else None
                    fat = row[8] if len(row) > 8 else None
                    mrg = row[9] if len(row) > 9 else None
                    d[cli.upper()] = {
                        'vol':    float(vol) if vol is not None else None,
                        'fat':    float(fat) if fat is not None else None,
                        'margem': float(mrg) if mrg is not None else 0.15,
                    }
                meta_global[ws_m.title.strip().upper()] = d
        except Exception:
            pass

    all_vendors = set(clientes_atual) | set(totais_atual)
    for p_dict in (ant_ano_h.get('vendedores', {}), ant_mes_h.get('vendedores', {})):
        all_vendors.update(p_dict.keys())
    all_vendors -= _EXCLUIDOS

    vendor_list = [v for v in VENDOR_ORDER if v in all_vendors]
    for v in sorted(all_vendors):
        if v not in vendor_list:
            vendor_list.append(v)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    geral_data = {}

    for vend in vendor_list:
        tab_name   = VENDOR_TAB.get(vend, vend.upper())
        title_name = VENDOR_TITLE.get(vend, tab_name)
        ws = wb.create_sheet(title=tab_name)

        cpp = {
            'ant_ano': ant_ano_h.get('vendedores', {}).get(vend, {}).get('clientes', {}),
            'ant_mes': ant_mes_h.get('vendedores', {}).get(vend, {}).get('clientes', {}),
            'atual':   clientes_atual.get(vend, {}),
        }
        tpp = {
            'ant_ano': ant_ano_h.get('vendedores', {}).get(vend, {}).get('TOTAL'),
            'ant_mes': ant_mes_h.get('vendedores', {}).get(vend, {}).get('TOTAL'),
            'atual':   totais_atual.get(vend),
        }
        meta_vend = meta_global.get(tab_name.upper(), {})

        _build_vendedor_sheet(ws, title_name, cpp, tpp, meta_vend, labels, ref_label)
        geral_data[vend] = {p: tpp[p] for p in _GERAL_PERS}

    ws_geral = wb.create_sheet(title='GERAL')
    _build_geral_sheet(ws_geral, geral_data, labels, ref_label)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
