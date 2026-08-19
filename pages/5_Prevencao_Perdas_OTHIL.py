"""Página Prevenção de Perdas — Estoque Parado."""
import datetime
import io
import json
import subprocess
import tempfile
import os

import streamlit as st
import pandas as pd

from parsers_estoque import parse_estoque_fisico
import data_store as ds
import comparativo

MODULO = 'prevencao_perdas'

_GERENCIA_DIR   = os.path.join(os.path.dirname(__file__), '..', 'gerencia_data')
_PREVPERDAS_DIR = os.path.join(_GERENCIA_DIR, 'prevencao_perdas')
_PERDAS_DIR     = os.path.join(_GERENCIA_DIR, 'perdas_realizadas')

st.set_page_config(page_title="Prevenção de Perdas", layout="wide")

st.title("🚨 Prevenção de Perdas — Estoque Parado")
st.caption("Identifica produtos parados antes que virem prejuízo.")

st.session_state.setdefault('usuario_nome', 'Ingrid')

# ── Constantes ────────────────────────────────────────────────────────────────
_NIVEL = {
    4: '🔴 Crítico',
    3: '🟠 Alta Prioridade',
    2: '🟡 Atenção',
    1: '🟢 Controlado',
}
_ACAO = {
    4: 'Promoção urgente / Vender urgente',
    3: 'Priorizar venda / Contatar vendedor',
    2: 'Monitorar / Ação preventiva',
    1: 'Manter monitoramento',
}

# ── Lógica de negócio ─────────────────────────────────────────────────────────

def _classificar(dias_estoque, valor_estoque, qtd_vendida):
    """Retorna nível de risco 1-4."""
    if qtd_vendida == 0:
        if dias_estoque > 30:   nivel = 4
        elif dias_estoque > 14: nivel = 3
        else:                   nivel = 2
    else:
        if dias_estoque > 60:   nivel = 4
        elif dias_estoque > 45: nivel = 3
        elif dias_estoque > 30: nivel = 2
        else:                   nivel = 1
    # Escalada por alto valor parado
    if valor_estoque > 10_000 and nivel < 4:
        nivel += 1
    return nivel


def _gerar_df(produtos, emissao_date, period_days):
    """Gera DataFrame gerencial completo para todos os produtos com saldo > 0."""
    rows = []
    for p in produtos:
        if p['saldo_atual'] <= 0:
            continue

        dias_est = (emissao_date - p['data_entrada']).days
        valor    = p['valor_estoque']
        qtd      = p['qtd_vendida']
        nivel    = _classificar(dias_est, valor, qtd)

        # Cobertura em dias
        if qtd > 0 and period_days > 0:
            cobertura = round(p['saldo_atual'] / (qtd / period_days))
        else:
            cobertura = None  # sem venda → sem cobertura calculável

        # Giro (% do estoque anterior consumido no período)
        ref  = p['anterior'] if p['anterior'] > 0 else max(p['saldo_atual'], 1)
        giro = round(qtd / ref * 100, 1) if ref > 0 else 0.0

        rows.append({
            '_nivel_num':            nivel,
            'Prioridade':            _NIVEL[nivel],
            'Produto':               p['produto'],
            'Responsável':           p['complemento'],
            'Código':                p['codigo'],
            'Data Entrada':          p['data_entrada_str'],
            'Dias em Estoque':       dias_est,
            'Dias sem Venda':        period_days if qtd == 0 else 0,
            'Estoque Atual (cx)':    p['saldo_atual'],
            'Qtd Vendida':           qtd,
            'Giro (%)':              giro,
            'Cobertura (dias)':      cobertura,
            'Custo Unit. (R$)':      p['custo_unit'],
            'Valor em Estoque (R$)': valor,
            'Margem Venda (R$)':     p['md_venda'],
            'Ação Recomendada':      _ACAO[nivel],
            'Observações':           '',
        })

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df = df.sort_values(
        ['_nivel_num', 'Valor em Estoque (R$)', 'Dias em Estoque'],
        ascending=[False, False, False],
    ).drop(columns=['_nivel_num']).reset_index(drop=True)
    df.index += 1
    return df


# ── Excel ─────────────────────────────────────────────────────────────────────

def _gerar_excel(df, titulo, emissao_str, periodo_str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = 'Prevenção de Perdas'

    # Paleta de cores OTHIL
    def fill(hex_): return PatternFill('solid', fgColor=hex_)

    COR = {
        'titulo':       fill('1B4332'),
        'resumo_hdr':   fill('2D6A4F'),
        'resumo_val':   fill('D8F3DC'),
        'cabecalho':    fill('40916C'),
        'critico':      fill('FFDADA'),
        'alta':         fill('FFE8CC'),
        'atencao':      fill('FFF9CC'),
        'controlado':   fill('D8F3DC'),
        'alt_critico':  fill('FFBABA'),
        'alt_alta':     fill('FFD5A8'),
        'alt_atencao':  fill('FFF3A8'),
        'alt_control':  fill('B7EBC9'),
    }
    NIVEL_FILL = {
        '🔴 Crítico':         (COR['critico'],    COR['alt_critico']),
        '🟠 Alta Prioridade': (COR['alta'],        COR['alt_alta']),
        '🟡 Atenção':         (COR['atencao'],     COR['alt_atencao']),
        '🟢 Controlado':      (COR['controlado'],  COR['alt_control']),
    }

    def fnt(bold=False, size=10, color='000000'):
        return Font(name='Arial', bold=bold, size=size, color=color)

    borda = Border(
        left=Side(style='thin', color='BBBBBB'),
        right=Side(style='thin', color='BBBBBB'),
        top=Side(style='thin', color='BBBBBB'),
        bottom=Side(style='thin', color='BBBBBB'),
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_a = Alignment(horizontal='left',   vertical='center')
    right_a= Alignment(horizontal='right',  vertical='center')

    def w(row, col, value, fill_=None, font_=None, align_=None, fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        if fill_:  c.fill  = fill_
        if font_:  c.font  = font_
        if align_: c.alignment = align_
        if fmt:    c.number_format = fmt
        c.border = borda
        return c

    N_COLS = 16  # número de colunas de dados

    # ── Linha 1: Título ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
    header_txt = f'OTHIL — {titulo}    |    Emissão: {emissao_str}'
    if periodo_str:
        header_txt += f'    |    Período desde: {periodo_str}'
    w(1, 1, header_txt, fill_=COR['titulo'], font_=fnt(bold=True, size=14, color='FFFFFF'), align_=center)
    ws.row_dimensions[1].height = 32

    # ── Linhas 2-4: Painel Resumo ─────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N_COLS)
    w(2, 1, 'PAINEL RESUMO EXECUTIVO',
      fill_=COR['resumo_hdr'], font_=fnt(bold=True, size=11, color='FFFFFF'), align_=center)
    ws.row_dimensions[2].height = 22

    total       = len(df)
    criticos    = int((df['Prioridade'] == '🔴 Crítico').sum())
    alta        = int((df['Prioridade'] == '🟠 Alta Prioridade').sum())
    valor_risco = float(df['Valor em Estoque (R$)'].sum())
    sem_v7      = int((df['Dias sem Venda'] >= 7).sum())
    est30       = int((df['Dias em Estoque'] >= 30).sum())

    cards = [
        ('Total Monitorado',   total),
        ('🔴 Críticos',        criticos),
        ('🟠 Alta Prioridade', alta),
        ('💰 Valor em Risco',  f"R$ {valor_risco:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')),
        ('Sem Venda ≥ 7 dias', sem_v7),
        ('Estoque ≥ 30 dias',  est30),
    ]
    # 6 cards, cada um ocupa ~2-3 colunas. Vamos usar 2 cols cada = 12 cols, centralizados
    col_starts = [1, 3, 5, 7, 10, 13]
    col_ends   = [2, 4, 6, 9, 12, N_COLS]
    for i, ((lbl, val), cs, ce) in enumerate(zip(cards, col_starts, col_ends)):
        ws.merge_cells(start_row=3, start_column=cs, end_row=3, end_column=ce)
        ws.merge_cells(start_row=4, start_column=cs, end_row=4, end_column=ce)
        w(3, cs, lbl, fill_=COR['resumo_hdr'], font_=fnt(bold=True, size=9, color='FFFFFF'), align_=center)
        w(4, cs, val, fill_=COR['resumo_val'], font_=fnt(bold=True, size=12, color='1B4332'), align_=center)
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 26

    # Linha 5: espaço
    ws.row_dimensions[5].height = 6

    # ── Linha 6: Cabeçalho da tabela ─────────────────────────────────────────
    HDR = 6
    COLUNAS = [
        'Prioridade', 'Produto', 'Responsável', 'Código',
        'Data Entrada', 'Dias em\nEstoque', 'Dias sem\nVenda',
        'Estoque\nAtual (cx)', 'Qtd\nVendida', 'Giro\n(%)',
        'Cobertura\n(dias)', 'Custo Unit.\n(R$)', 'Valor em\nEstoque (R$)',
        'Margem\nVenda (R$)', 'Ação Recomendada', 'Observações',
    ]
    for c, nome in enumerate(COLUNAS, start=1):
        w(HDR, c, nome, fill_=COR['cabecalho'],
          font_=fnt(bold=True, size=9, color='FFFFFF'), align_=center)
    ws.row_dimensions[HDR].height = 36

    # ── Dados ─────────────────────────────────────────────────────────────────
    FMT_CX   = '#,##0.000'
    FMT_MON  = 'R$ #,##0.00'
    FMT_PCT  = '0.0%'
    FMT_INT  = '0'

    nivel_count = {}  # para alternar cor por bloco de mesmo nível

    for r_idx, row in df.iterrows():
        xl_row = HDR + r_idx   # r_idx começa em 1
        nivel  = row['Prioridade']
        nivel_count[nivel] = nivel_count.get(nivel, 0) + 1
        pares = NIVEL_FILL.get(nivel, (COR['controlado'], COR['alt_control']))
        fill_ = pares[nivel_count[nivel] % 2]

        giro_val = row['Giro (%)']
        giro_frac = giro_val / 100 if giro_val is not None else None

        vals_fmts_aligns = [
            (row['Prioridade'],            None,    center),
            (row['Produto'],               None,    left_a),
            (row['Responsável'],           None,    center),
            (row['Código'],                None,    center),
            (row['Data Entrada'],          None,    center),
            (row['Dias em Estoque'],       FMT_INT, right_a),
            (row['Dias sem Venda'],        FMT_INT, right_a),
            (row['Estoque Atual (cx)'],    FMT_CX,  right_a),
            (row['Qtd Vendida'],           FMT_CX,  right_a),
            (giro_frac,                    FMT_PCT, right_a),
            (row['Cobertura (dias)'],      FMT_INT, right_a),
            (row['Custo Unit. (R$)'],      FMT_MON, right_a),
            (row['Valor em Estoque (R$)'], FMT_MON, right_a),
            (row['Margem Venda (R$)'],     FMT_MON, right_a),
            (row['Ação Recomendada'],      None,    left_a),
            (row['Observações'],           None,    left_a),
        ]
        for c_idx, (val, fmt, aln) in enumerate(vals_fmts_aligns, start=1):
            w(xl_row, c_idx, val, fill_=fill_, font_=fnt(size=9), align_=aln, fmt=fmt)
        ws.row_dimensions[xl_row].height = 16

    # ── Larguras das colunas ──────────────────────────────────────────────────
    widths = [22, 38, 14, 14, 13, 10, 10, 12, 10, 8, 10, 14, 18, 14, 38, 28]
    for i, wid in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wid

    # Congelar a partir da linha 7 (deixa cabeçalho fixo)
    ws.freeze_panes = 'A7'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── UI Helpers ────────────────────────────────────────────────────────────────

def _publicar_gerencia(df, emissao_str, periodo_str, tipo):
    """Salva snapshot JSON na pasta gerencia_data/prevencao_perdas/."""
    os.makedirs(_PREVPERDAS_DIR, exist_ok=True)
    slug = f"{datetime.date.today().strftime('%Y-%m-%d')}_{tipo}"
    snapshot = {
        'publicado_em': datetime.datetime.now().isoformat(),
        'emissao':      emissao_str,
        'periodo':      periodo_str or '',
        'tipo':         tipo,
        'resumo': {
            'total':           len(df),
            'criticos':        int((df['Prioridade'] == '🔴 Crítico').sum()),
            'alta_prioridade': int((df['Prioridade'] == '🟠 Alta Prioridade').sum()),
            'atencao':         int((df['Prioridade'] == '🟡 Atenção').sum()),
            'valor_risco':     float(df['Valor em Estoque (R$)'].sum()),
            'sem_venda_7d':    int((df['Dias sem Venda'] >= 7).sum()),
            'estoque_30d':     int((df['Dias em Estoque'] >= 30).sum()),
        },
        'produtos': [
            {
                'prioridade':    r['Prioridade'],
                'produto':       r['Produto'],
                'responsavel':   r['Responsável'],
                'dias_estoque':  int(r['Dias em Estoque']),
                'dias_sem_venda': int(r['Dias sem Venda']),
                'saldo_cx':      float(r['Estoque Atual (cx)']),
                'qtd_vendida':   float(r['Qtd Vendida']),
                'valor_estoque': float(r['Valor em Estoque (R$)']),
                'acao':          r['Ação Recomendada'],
            }
            for _, r in df.iterrows()
        ],
    }
    path = os.path.join(_PREVPERDAS_DIR, f'{slug}.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(snapshot, f, ensure_ascii=False, indent=2)
    # Persistência real e versionada (sobrevive a restart do Streamlit Cloud;
    # antes só existia em gerencia_data/, apagado a cada restart)
    try:
        ds.save_record(
            modulo=MODULO, tipo_periodo='diario', periodo_ref=slug,
            valores=snapshot, usuario=st.session_state.get('usuario_nome'),
        )
    except Exception as e:
        st.warning(f'Publicado localmente, mas houve um problema ao salvar de forma permanente: {e}')
    return slug


def _pdf_to_text(uploaded_file):
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
        tmp.write(uploaded_file.read())
        tmp_path = tmp.name
    try:
        result = subprocess.run(
            ['pdftotext', '-layout', tmp_path, '-'],
            capture_output=True, text=True, timeout=30,
        )
        return result.stdout if result.returncode == 0 else None
    except Exception:
        return None
    finally:
        os.unlink(tmp_path)


def _publicacao_anterior(tipo: str):
    """Última publicação salva para este tipo (sem_venda / mes_estoque),
    usada como base do comparativo — não inclui os dados que ainda estão
    só na tela (não publicados)."""
    try:
        slugs = [s for s in ds.list_periodos(MODULO, 'diario') if tipo in s]
        if not slugs:
            return None
        return ds.load_current(MODULO, 'diario', slugs[0])
    except Exception:
        return None


def _cards(df, tipo: str = None):
    total       = len(df)
    criticos    = int((df['Prioridade'] == '🔴 Crítico').sum())
    alta        = int((df['Prioridade'] == '🟠 Alta Prioridade').sum())
    valor_risco = float(df['Valor em Estoque (R$)'].sum())
    sem_v7      = int((df['Dias sem Venda'] >= 7).sum())
    est30       = int((df['Dias em Estoque'] >= 30).sum())

    vr_fmt = f"R$ {valor_risco:,.0f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("📦 Total monitorado",    total)
    c2.metric("🔴 Críticos",            criticos)
    c3.metric("🟠 Alta Prioridade",     alta)
    c4.metric("💰 Valor em risco",      vr_fmt)
    c5.metric("🕐 Sem venda ≥ 7 dias",  sem_v7)
    c6.metric("📅 Estoque ≥ 30 dias",   est30)

    if tipo:
        anterior = _publicacao_anterior(tipo)
        if anterior:
            resumo_ant = anterior.get('valores', {}).get('resumo', {})
            st.caption(
                f"📊 Comparativo vs última publicação "
                f"({anterior.get('atualizado_em', '')[:10]}):"
            )
            pc1, pc2, pc3 = st.columns(3)
            comp_total = comparativo.calcular(total, resumo_ant.get('total'), menor_e_melhor=True)
            comp_crit  = comparativo.calcular(criticos, resumo_ant.get('criticos'), menor_e_melhor=True)
            comp_vr    = comparativo.calcular(valor_risco, resumo_ant.get('valor_risco'), menor_e_melhor=True)
            pc1.metric("📦 Total", total, delta=comparativo.formatar_variacao(comp_total, casas=1))
            pc2.metric("🔴 Críticos", criticos, delta=comparativo.formatar_variacao(comp_crit, casas=1))
            pc3.metric("💰 Valor em risco", vr_fmt, delta=comparativo.formatar_variacao(comp_vr, casas=1))


def _col_config():
    return {
        'Estoque Atual (cx)':    st.column_config.NumberColumn(format='%.3f'),
        'Qtd Vendida':           st.column_config.NumberColumn(format='%.3f'),
        'Giro (%)':              st.column_config.NumberColumn(format='%.1f%%'),
        'Cobertura (dias)':      st.column_config.NumberColumn(format='%d'),
        'Dias em Estoque':       st.column_config.NumberColumn(format='%d'),
        'Dias sem Venda':        st.column_config.NumberColumn(format='%d'),
        'Custo Unit. (R$)':      st.column_config.NumberColumn(format='R$ %.2f'),
        'Valor em Estoque (R$)': st.column_config.NumberColumn(format='R$ %.2f'),
        'Margem Venda (R$)':     st.column_config.NumberColumn(format='R$ %.2f'),
    }


def _render_tab(key, titulo, upload_label, filtro_sem_venda, dias_min):
    uploaded = st.file_uploader(upload_label, type='pdf', key=f'up_{key}')
    if not uploaded:
        return

    with st.spinner("Processando PDF..."):
        texto = _pdf_to_text(uploaded)
    if not texto:
        st.error("Não foi possível processar o PDF.")
        return

    dados        = parse_estoque_fisico(texto)
    emissao_str  = dados['emissao']
    emissao_date = dados['emissao_date']
    periodo_str  = dados['periodo']
    produtos     = dados['produtos']

    if periodo_str:
        try:
            pd_ = datetime.datetime.strptime(periodo_str, '%d/%m/%Y').date()
            period_days = max((emissao_date - pd_).days, 1)
        except Exception:
            period_days = 7
    else:
        period_days = 1

    st.success(
        f"PDF lido · Emissão: **{emissao_str}**"
        + (f" · Período desde: **{periodo_str}**" if periodo_str else "")
        + f" · Total no relatório: **{len(produtos)} produtos**"
    )

    # Gerar DataFrame completo e aplicar filtros de negócio
    df_full = _gerar_df(produtos, emissao_date, period_days)
    if df_full.empty:
        st.success("✅ Nenhum produto encontrado.")
        return

    if filtro_sem_venda:
        df_full = df_full[
            (df_full['Qtd Vendida'] == 0) & (df_full['Dias em Estoque'] > 0)
        ].reset_index(drop=True)
        df_full.index += 1
    if dias_min > 0:
        df_full = df_full[df_full['Dias em Estoque'] >= dias_min].reset_index(drop=True)
        df_full.index += 1

    if df_full.empty:
        st.success("✅ Nenhum produto encontrado com os critérios definidos.")
        return

    # Painel resumo
    st.markdown("---")
    _cards(df_full, tipo=key)
    st.markdown("---")

    # Filtros
    with st.expander("🔍 Pesquisa e Filtros", expanded=True):
        sc1, sc2 = st.columns(2)
        busca_prod = sc1.text_input("Buscar produto", key=f'bp_{key}').strip().upper()
        busca_resp = sc2.text_input("Buscar responsável / depto", key=f'br_{key}').strip().upper()

        st.markdown("**Filtros rápidos:**")
        fc1, fc2, fc3, fc4 = st.columns(4)
        f_crit  = fc1.checkbox("🔴 Só Críticos",          key=f'f1_{key}')
        f_alta  = fc2.checkbox("🟠 Críticos + Alta Prior.", key=f'f2_{key}')
        f_sv    = fc3.checkbox("Sem venda no período",      key=f'f3_{key}')
        f_30d   = fc4.checkbox("≥ 30 dias em estoque",      key=f'f4_{key}')

    # Aplicar filtros interativos
    df_show = df_full.copy()
    if busca_prod:
        df_show = df_show[df_show['Produto'].str.contains(busca_prod, na=False)]
    if busca_resp:
        df_show = df_show[df_show['Responsável'].str.contains(busca_resp, na=False)]
    if f_crit:
        df_show = df_show[df_show['Prioridade'] == '🔴 Crítico']
    if f_alta:
        df_show = df_show[df_show['Prioridade'].isin(['🔴 Crítico', '🟠 Alta Prioridade'])]
    if f_sv:
        df_show = df_show[df_show['Qtd Vendida'] == 0]
    if f_30d:
        df_show = df_show[df_show['Dias em Estoque'] >= 30]

    if df_show.empty:
        st.info("Nenhum produto corresponde aos filtros selecionados.")
        return

    # Colunas visíveis na tela (enxuto)
    cols_tela = [
        'Prioridade', 'Produto', 'Responsável',
        'Dias em Estoque', 'Dias sem Venda',
        'Estoque Atual (cx)', 'Qtd Vendida', 'Giro (%)',
        'Cobertura (dias)', 'Valor em Estoque (R$)', 'Ação Recomendada',
    ]
    st.dataframe(df_show[cols_tela], use_container_width=True, column_config=_col_config())
    st.caption(f"Exibindo **{len(df_show)}** produto(s). O Excel inclui todas as colunas (Código, Custo, Observações…).")

    # Download Excel
    excel = _gerar_excel(df_show, titulo, emissao_str, periodo_str or '')
    col_dl, col_pub = st.columns([2, 1])
    with col_dl:
        if excel:
            nome_arquivo = f"prevencao_perdas_{key}_{emissao_str.replace('/', '-')}.xlsx"
            st.download_button(
                "⬇️ Baixar Planilha Excel",
                data=excel,
                file_name=nome_arquivo,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                type='primary',
                use_container_width=True,
            )
        else:
            csv = df_show.to_csv(sep=';', decimal=',').encode('utf-8-sig')
            st.download_button("⬇️ Baixar CSV", data=csv,
                               file_name=f"prevperdas_{key}.csv", mime='text/csv',
                               use_container_width=True)
    with col_pub:
        if st.button("📤 Publicar na Gerência", key=f'pub_{key}', use_container_width=True):
            slug = _publicar_gerencia(df_show, emissao_str, periodo_str or '', key)
            st.success(f"✅ Publicado! Disponível na Gerência como **{slug}**.")


# ── Alerta de Recebimento ─────────────────────────────────────────────────────

def _carregar_pp_mais_recente():
    """Retorna o snapshot de PP mais recente publicado (qualquer tipo).
    Lê da persistência real (data_store) primeiro — sobrevive a restart do
    app; arquivo local entra como complemento/fallback."""
    try:
        slugs = ds.list_periodos(MODULO, 'diario')
        if slugs:
            registro = ds.load_current(MODULO, 'diario', slugs[0])
            if registro:
                return registro['valores']
    except Exception:
        pass

    if not os.path.isdir(_PREVPERDAS_DIR):
        return None
    arquivos = sorted([f for f in os.listdir(_PREVPERDAS_DIR) if f.endswith('.json')], reverse=True)
    if not arquivos:
        return None
    try:
        with open(os.path.join(_PREVPERDAS_DIR, arquivos[0]), 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None


def _render_alerta_recebimento():
    st.subheader("🔔 Alerta de Recebimento")
    st.info(
        "Envie o **Estoque Físico do dia**. O app identifica produtos que "
        "**entraram hoje** e que **já estão na lista de estoque parado** — "
        "sinal de que está sendo comprado mais do que se consegue girar."
    )

    snap = _carregar_pp_mais_recente()
    if not snap:
        st.warning("Nenhum dado de Prevenção de Perdas publicado ainda. "
                   "Publique primeiro nas abas 1 ou 2.")
        return

    nomes_pp = {p['produto'].strip().upper() for p in snap.get('produtos', [])}
    tipo_pp  = '1 Semana Sem Venda' if snap.get('tipo') == 'sem_venda' else '1 Mês em Estoque'
    pub_em   = snap.get('publicado_em', '')[:10]
    st.caption(f"Comparando com: snapshot **{tipo_pp}** publicado em **{pub_em}** "
               f"({len(nomes_pp)} produtos em alerta)")

    uploaded = st.file_uploader("Estoque Físico — Do dia atual (PDF)",
                                 type='pdf', key='up_receb')
    if not uploaded:
        return

    with st.spinner("Processando PDF..."):
        texto = _pdf_to_text(uploaded)
    if not texto:
        st.error("Não foi possível processar o PDF.")
        return

    dados        = parse_estoque_fisico(texto)
    emissao_date = dados['emissao_date']
    emissao_str  = dados['emissao']
    produtos     = dados['produtos']

    recebidos = [p for p in produtos if p['data_entrada'] == emissao_date]

    st.success(
        f"PDF lido · Emissão: **{emissao_str}** · "
        f"**{len(recebidos)}** produto(s) com entrada hoje"
    )

    if not recebidos:
        st.info("Nenhum produto com entrada na data de emissão.")
        return

    alertas = [p for p in recebidos if p['produto'].strip().upper() in nomes_pp]
    sem_alerta = [p for p in recebidos if p not in alertas]

    if not alertas:
        st.success(f"✅ Nenhum dos {len(recebidos)} produtos recebidos hoje "
                   "consta na lista de estoque parado.")
    else:
        st.error(f"⚠️ **{len(alertas)} produto(s) recebido(s) hoje JÁ ESTÃO "
                 "na lista de estoque parado!**")
        df_al = pd.DataFrame([{
            'Produto':              p['produto'],
            'Responsável':         p['complemento'],
            'Saldo em Estoque (cx)': p['saldo_atual'],
            'Custo Unit. (R$)':    p['custo_unit'],
            'Valor Parado (R$)':   p['valor_estoque'],
        } for p in alertas])
        st.dataframe(df_al, use_container_width=True, hide_index=True,
                     column_config={
                         'Saldo em Estoque (cx)': st.column_config.NumberColumn(format='%.3f'),
                         'Custo Unit. (R$)':      st.column_config.NumberColumn(format='R$ %.2f'),
                         'Valor Parado (R$)':     st.column_config.NumberColumn(format='R$ %.2f'),
                     })

    if sem_alerta:
        with st.expander(f"Ver {len(sem_alerta)} produto(s) recebidos hoje sem alerta"):
            df_sa = pd.DataFrame([{
                'Produto':      p['produto'],
                'Responsável':  p['complemento'],
                'Saldo (cx)':   p['saldo_atual'],
            } for p in sem_alerta])
            st.dataframe(df_sa, use_container_width=True, hide_index=True,
                         column_config={'Saldo (cx)': st.column_config.NumberColumn(format='%.3f')})


# ── Tabs ──────────────────────────────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs([
    "🕐 1 Semana Sem Venda",
    "📦 1 Mês em Estoque",
    "🔔 Alerta de Recebimento",
])

with tab1:
    st.subheader("1 Semana Sem Venda")
    st.info(
        "Envie o Estoque Físico com **período de 1 semana**. "
        "Filtra produtos com **Saldo > 0 e Qtde Vendida = 0** no período."
    )
    _render_tab(
        key='sem_venda',
        titulo='Prevenção de Perdas — 1 Semana Sem Venda',
        upload_label='Estoque Físico — Período de 1 semana (PDF)',
        filtro_sem_venda=True,
        dias_min=0,
    )

with tab2:
    st.subheader("1 Mês em Estoque")
    st.info(
        "Envie o Estoque Físico **do dia atual**. "
        "Filtra produtos com **Saldo > 0 e Data de Entrada há mais de 30 dias**."
    )
    _render_tab(
        key='mes_estoque',
        titulo='Prevenção de Perdas — 1 Mês em Estoque',
        upload_label='Estoque Físico — Do dia atual (PDF)',
        filtro_sem_venda=False,
        dias_min=30,
    )

with tab3:
    _render_alerta_recebimento()
