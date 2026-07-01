"""Gera o Excel 'Relatorio_Diario_DD-MM-AAAA_OTHIL.xlsx' a partir do resultado
de parsers_diario.parse_relatorio_diario().

Dois indicadores de margem:
  MC R$  = Faturamento − Custo_PDF  (custo bruto sem ajuste)
  MC %   = MC_R$ / Custo_PDF × 100

  Resultado Real % = Resultado%_PDF + 15pp
    (o PDF calcula Resultado% = Resultado_R$ / Custo_PDF × 100;
     somamos 15pp para refletir a despesa administrativa embutida)

  Alerta / Impacto filtram itens com Resultado Real % < −15%.

Estrutura:
  - Leia-me
  - Resumo  (KPIs + ranking de vendedores + gráfico)
  - Alertas_<-15%  (itens com Resultado Real % < −15%)
  - Vendedor_<Nome>  (detalhado por produto + gerencial por cliente)
"""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

VERDE_BG,   VERDE_FG   = 'D8EFE3', '1B4332'
AMARELO_BG, AMARELO_FG = 'FEF9C3', '7D6608'
VERMELHO_BG,VERMELHO_FG= 'FADADD', '7A1F2B'
HEADER_BG,  HEADER_FG  = '2D6A4F', 'FFFFFF'
FONT_NAME = 'Arial'
MONEY_FMT = '#,##0.00'
PCT_FMT   = '0.00'

THIN   = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# Cálculo de margens
# ---------------------------------------------------------------------------

def _calc(faturamento, custo_pdf):
    """Retorna (mc_rs, mc_pct, resultado_real_pct).

    MC R$  = Fat − Custo_PDF
    MC %   = MC_R$ / Custo_PDF × 100
    Resultado Real % = MC% + 15  (Resultado%_PDF + 15pp de despesa adm)
    """
    mc_rs  = faturamento - custo_pdf
    mc_pct = mc_rs / custo_pdf * 100 if custo_pdf else 0.0
    resultado_real_pct = mc_pct + 15
    return round(mc_rs, 2), round(mc_pct, 2), round(resultado_real_pct, 2)


# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------

def _font(bold=False, size=11, color='000000'):
    return Font(name=FONT_NAME, bold=bold, size=size, color=color)


def _header_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = _font(bold=True, color=HEADER_FG)
    c.fill      = PatternFill('solid', fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border    = BORDER
    return c


def _mc_colors(pct):
    if pct >= 15: return VERDE_BG,    VERDE_FG
    if pct >= 0:  return AMARELO_BG,  AMARELO_FG
    return             VERMELHO_BG, VERMELHO_FG


def _pct_cell(ws, row, col, value, pct_for_color):
    cell = ws.cell(row=row, column=col, value=round(value, 2) if isinstance(value, float) else value)
    cell.number_format = PCT_FMT
    bg, fg = _mc_colors(pct_for_color)
    cell.fill   = PatternFill('solid', fgColor=bg)
    cell.font   = _font(color=fg)
    cell.border = BORDER
    return cell


def _money(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = MONEY_FMT
    cell.border = BORDER
    cell.font   = _font()
    return cell


def _qty(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = '#,##0.000'
    cell.border = BORDER
    cell.font   = _font()
    return cell


def _text(ws, row, col, value, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font   = _font(bold=bold)
    cell.border = BORDER
    return cell


def _sheet_name_for_vendor(nome):
    safe = re.sub(r'[\\/*?:\[\]]', '', nome)[:25]
    return f'Vendedor_{safe}'


# ---------------------------------------------------------------------------
# Aba "Dados" (oculta)
# Colunas: A Vendedor | B Vendedor_raw | C Cliente_cod | D Cliente_nome |
#          E Produto | F Qtd | G Fat | H CustoUnit_PDF | I CustoTotal_PDF |
#          J MC_RS | K MC_pct | L Resultado_Real_pct |
#          M Flag1stCliVend | N Flag1stCliGlobal
# ---------------------------------------------------------------------------

def _build_dados(wb, itens):
    ws = wb.create_sheet('Dados')
    headers = ['Vendedor', 'Vendedor_raw', 'Cliente_codigo', 'Cliente_nome',
               'Produto', 'Qtd', 'Faturamento', 'CustoUnit_PDF', 'CustoTotal_PDF',
               'MC_RS', 'MC_pct', 'Resultado_Real_pct', 'Flag1stCliVend', 'Flag1stCliGlobal']
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h).font = _font(bold=True)

    n = len(itens)
    for i, it in enumerate(itens):
        r = i + 2
        mc_rs, mc_pct, res_real = _calc(it['faturamento'], it['custo_total'])
        ws.cell(row=r, column=1,  value=it['vendedor'] or it['vendedor_raw'])
        ws.cell(row=r, column=2,  value=it['vendedor_raw'])
        ws.cell(row=r, column=3,  value=it['cliente_codigo'])
        ws.cell(row=r, column=4,  value=it['cliente_nome'])
        ws.cell(row=r, column=5,  value=it['produto'])
        ws.cell(row=r, column=6,  value=it['qtd'])
        ws.cell(row=r, column=7,  value=it['faturamento'])
        ws.cell(row=r, column=8,  value=it['custo_unit'])
        ws.cell(row=r, column=9,  value=it['custo_total'])
        ws.cell(row=r, column=10, value=round(mc_rs,   2))
        ws.cell(row=r, column=11, value=round(mc_pct,  4))
        ws.cell(row=r, column=12, value=round(res_real, 4))
        ws.cell(row=r, column=13, value=f'=IF(COUNTIFS($B$2:B{r},B{r},$C$2:C{r},C{r})=1,1,0)')
        ws.cell(row=r, column=14, value=f'=IF(COUNTIFS($C$2:C{r},C{r})=1,1,0)')
    ws.sheet_state = 'hidden'
    return ws, n


# ---------------------------------------------------------------------------
# Aba "Leia-me"
# ---------------------------------------------------------------------------

def _build_leiame(wb, data_emissao, periodo):
    ws = wb.create_sheet('Leia-me')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 110
    row = [1]

    def add(text, bold=False, size=11, color='000000', fill=None):
        r = row[0]
        c = ws.cell(row=r, column=1, value=text)
        c.font      = _font(bold=bold, size=size, color=color)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if fill:
            c.fill = PatternFill('solid', fgColor=fill)
        row[0] += 1
        return c

    add('Relatório Diário de Vendas OTHIL', bold=True, size=18, color=HEADER_BG)
    add(f'Dia: {data_emissao or "-"}    Período: {periodo or "-"}', bold=True, size=12)
    add('')
    add('O QUE É ESTE ARQUIVO', bold=True, size=13, color=HEADER_BG)
    add('Gerado automaticamente a partir do PDF "Lucratividade por Vendedor-Cliente '
        'no Previsão" (Mercatus) do dia. Cada linha de produto do PDF é extraída '
        'item a item (produto x cliente x vendedor).')
    add('')
    add('INDICADORES DE MARGEM', bold=True, size=13, color=HEADER_BG)
    add('MC R$ = Faturamento − Custo_PDF  (custo bruto do sistema, sem ajuste)')
    add('MC %  = MC_R$ / Custo_PDF × 100  (margem de contribuição sobre o custo)')
    add('Resultado Real % = MC% + 15pp  '
        '(o sistema já embute 15% de despesa administrativa no custo; '
        'somamos 15pp para refletir o resultado real)')
    add('')
    add('COMO LER CADA ABA', bold=True, size=13, color=HEADER_BG)
    add('Resumo: KPIs principais do dia e ranking de vendedores.', bold=True)
    add('Alertas_<-15%: itens com Resultado Real % abaixo de −15% (pior primeiro). '
        'Mostra Qtd, Custo Unit., Venda Unit., MC % e Resultado Real %.', bold=True)
    add('Vendedor_<Nome>: detalhamento por produto (esquerda) e por cliente (direita).', bold=True)
    add('')
    add('SEMÁFORO DE CORES', bold=True, size=13, color=HEADER_BG)
    add('Verde   — % ≥ 15%  (saudável)',          color=VERDE_FG,    fill=VERDE_BG)
    add('Amarelo — % entre 0% e 15%  (atenção)',  color=AMARELO_FG,  fill=AMARELO_BG)
    add('Vermelho — % negativa  (prejuízo)',       color=VERMELHO_FG, fill=VERMELHO_BG)
    return ws


# ---------------------------------------------------------------------------
# Aba "Resumo"
# ---------------------------------------------------------------------------

def _build_resumo(wb, n_dados, vendor_order):
    ws = wb.create_sheet('Resumo')
    ws.sheet_view.showGridLines = False

    def dr(col):   # range na aba Dados
        return f'Dados!${col}$2:${col}${n_dados + 1}'

    # KPIs
    kpis = [
        ('Faturamento',       f'=SUM({dr("G")})',                                                  MONEY_FMT),
        ('MC R$',             f'=SUM({dr("J")})',                                                  MONEY_FMT),
        ('Resultado Real %',  f'=IFERROR(SUM({dr("J")})/SUM({dr("I")})*100+15,0)',                PCT_FMT),
        ('Caixas',            f'=SUM({dr("F")})',                                                  '#,##0.000'),
        ('Clientes',          f'=SUM({dr("N")})',                                                   '0'),
        ('Vendedores Ativos', f'={len(vendor_order)}',                                             '0'),
    ]
    ws.cell(row=1, column=1, value='Relatório Diário de Vendas OTHIL').font = _font(bold=True, size=18, color=HEADER_BG)
    col = 1
    for label, formula, fmt in kpis:
        lc = ws.cell(row=4, column=col, value=label)
        lc.font      = _font(bold=True, color=HEADER_FG)
        lc.fill      = PatternFill('solid', fgColor=HEADER_BG)
        lc.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        vc = ws.cell(row=5, column=col, value=formula)
        vc.font          = _font(bold=True, size=14)
        vc.number_format = fmt
        vc.alignment     = Alignment(horizontal='center')
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        for rr in (4, 5):
            for cc in (col, col + 1):
                ws.cell(row=rr, column=cc).border = BORDER
        col += 2

    # Ranking de vendedores
    rank_row0 = 8
    headers = ['Vendedor', 'Clientes', 'Caixas', 'Faturamento R$', 'Fat. Unit. R$',
               'Custo Unit. R$', 'MC R$', 'Resultado Real %',
               'Ticket Médio', 'Itens MC<-15%', 'Status']
    for j, h in enumerate(headers, start=1):
        _header_cell(ws, rank_row0, j, h)

    b  = dr('A')   # Vendedor
    f_ = dr('F')   # Qtd
    g  = dr('G')   # Faturamento
    i_ = dr('I')   # CustoTotal_PDF
    j_ = dr('J')   # MC_RS
    k  = dr('K')   # MC_pct
    m_ = dr('M')   # Flag1stCliVend

    for i, item in enumerate(vendor_order):
        vname, mc_pct_hint, res_real_hint = item
        r = rank_row0 + 1 + i
        crit = f'{b},"{vname}"'
        _text(ws, r, 1, vname, bold=True)
        c2 = ws.cell(row=r, column=2, value=f'=SUMIFS({m_},{crit})')
        c2.border = BORDER; c2.font = _font()
        _qty  (ws, r, 3, f'=SUMIFS({f_},{crit})')
        _money(ws, r, 4, f'=SUMIFS({g},{crit})')
        _money(ws, r, 5, f'=IFERROR(D{r}/C{r},0)')
        _money(ws, r, 6, f'=IFERROR(SUMIFS({i_},{crit})/C{r},0)')
        _money(ws, r, 7, f'=SUMIFS({j_},{crit})')                              # MC R$
        _pct_cell(ws, r, 8, f'=IFERROR(H{r}/SUMIFS({i_},{crit})*100+15,0)', res_real_hint)  # Resultado Real %
        _money(ws, r, 9, f'=IFERROR(D{r}/B{r},0)')                            # Ticket Médio
        c10 = ws.cell(row=r, column=10, value=f'=COUNTIFS({b},"{vname}",{k},"<-15")')
        c10.border = BORDER; c10.font = _font()
        sc = ws.cell(row=r, column=11,
                     value=f'=IF(H{r}>=15,"OK",IF(H{r}>=0,"Atenção","Crítico"))')
        sc.font = _font(bold=True); sc.alignment = Alignment(horizontal='center'); sc.border = BORDER

    last_row = rank_row0 + len(vendor_order)
    for j in range(1, 12):
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.column_dimensions['A'].width = 14

    chart = BarChart()
    chart.title = 'Faturamento por Vendedor'
    chart.y_axis.title = 'R$'; chart.x_axis.title = 'Vendedor'; chart.style = 10
    data = Reference(ws, min_col=4, min_row=rank_row0, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=rank_row0 + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height, chart.width = 10, 24
    ws.add_chart(chart, f'A{last_row + 3}')
    return ws


# ---------------------------------------------------------------------------
# Aba "Alertas_<-15%"
# Filtro: Resultado Real % < -15%
# Colunas: Vendedor | Cliente | Produto | Qtd | Custo Unit. R$ | Venda Unit. R$ | MC % | Resultado Real %
# ---------------------------------------------------------------------------

def _build_alertas(wb, itens):
    ws = wb.create_sheet('Alertas_<-15%')
    ws.sheet_view.showGridLines = False

    headers = ['Vendedor', 'Cliente', 'Produto', 'Qtd',
               'Custo Unit. R$', 'Venda Unit. R$', 'MC R$', 'Resultado Real %']
    ws.row_dimensions[1].height = 32
    for j, h in enumerate(headers, start=1):
        _header_cell(ws, 1, j, h)

    rows = []
    for it in itens:
        mc_rs, mc_pct, res_real = _calc(it['faturamento'], it['custo_total'])
        if mc_pct < -15:           # filtro: MC % < -15%
            rows.append((it, mc_rs, res_real))
    rows.sort(key=lambda t: t[2])  # pior Resultado Real % primeiro

    for i, (it, mc_rs, res_real) in enumerate(rows):
        r = i + 2
        qtd        = it['qtd']
        venda_unit = it['faturamento'] / qtd if qtd else 0.0
        _text (ws, r, 1, it['vendedor'] or it['vendedor_raw'])
        _text (ws, r, 2, it['cliente_nome'])
        _text (ws, r, 3, it['produto'])
        _qty  (ws, r, 4, qtd)
        _money(ws, r, 5, it['custo_unit'])
        _money(ws, r, 6, round(venda_unit, 2))
        _money(ws, r, 7, round(mc_rs, 2))
        _pct_cell(ws, r, 8, round(res_real, 2), res_real)

    widths = [14, 38, 42, 10, 14, 14, 14, 16]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A2'
    return ws, len(rows)


# ---------------------------------------------------------------------------
# Aba "Vendedor_<Nome>"
# Esquerda: Produto por produto
# Direita:  Gerencial por cliente
# ---------------------------------------------------------------------------

def _build_vendedor_sheet(wb, vname, vendor_itens):
    ws = wb.create_sheet(_sheet_name_for_vendor(vname))
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value=f'Vendedor: {vname}').font = _font(bold=True, size=16, color=HEADER_BG)

    # --- BLOCO ESQUERDO: detalhado por produto ---
    # A=Cliente | B=Produto | C=Qtd | D=Fat R$ | E=Fat Unit R$ |
    # F=Custo Unit R$ | G=MC R$ | H=Resultado Real %
    LEFT_HEADERS = ['Cliente', 'Produto', 'Qtd', 'Fat. R$', 'Fat. Unit. R$',
                    'Custo Unit. R$', 'MC R$', 'Resultado Real %']
    title_row  = 3
    header_row = 4
    ws.cell(row=title_row, column=1, value='DETALHADO POR PRODUTO').font = _font(bold=True, size=12, color=HEADER_BG)
    for j, h in enumerate(LEFT_HEADERS, start=1):
        _header_cell(ws, header_row, j, h)

    first_data_row = header_row + 1
    for i, it in enumerate(vendor_itens):
        r = first_data_row + i
        mc_rs, mc_pct, res_real = _calc(it['faturamento'], it['custo_total'])
        fat_unit = it['faturamento'] / it['qtd'] if it['qtd'] else 0.0
        _text    (ws, r, 1, it['cliente_nome'])
        _text    (ws, r, 2, it['produto'])
        _qty     (ws, r, 3, it['qtd'])
        _money   (ws, r, 4, it['faturamento'])
        _money   (ws, r, 5, round(fat_unit, 2))
        _money   (ws, r, 6, it['custo_unit'])
        _money   (ws, r, 7, round(mc_rs,  2))
        _pct_cell(ws, r, 8, round(res_real, 2), res_real)

    # Linha TOTAL
    left_total_row = first_data_row + len(vendor_itens)
    total_fat   = sum(it['faturamento'] for it in vendor_itens)
    total_custo = sum(it['custo_total'] for it in vendor_itens)
    total_qtd   = sum(it['qtd']         for it in vendor_itens)
    total_mc_rs, total_mc_pct, total_res_real = _calc(total_fat, total_custo)
    avg_custo_unit = total_custo / total_qtd if total_qtd else 0.0
    avg_fat_unit   = total_fat   / total_qtd if total_qtd else 0.0

    _text    (ws, left_total_row, 1, 'TOTAL', bold=True)
    ws.cell(row=left_total_row, column=1).fill = PatternFill('solid', fgColor='E9F5EF')
    _qty     (ws, left_total_row, 3, total_qtd)
    _money   (ws, left_total_row, 4, round(total_fat, 2))
    _money   (ws, left_total_row, 5, round(avg_fat_unit, 2))
    _money   (ws, left_total_row, 6, round(avg_custo_unit, 2))
    _money   (ws, left_total_row, 7, round(total_mc_rs,  2))
    _pct_cell(ws, left_total_row, 8, round(total_res_real, 2), total_res_real)
    ws.cell(row=left_total_row, column=1).font = _font(bold=True)

    # --- BLOCO DIREITO: gerencial por cliente ---
    RIGHT_COL0 = 10   # coluna J
    RIGHT_HEADERS = ['Cliente', 'Qtd', 'Fat. R$', 'Fat. Unit. R$',
                     'Custo Unit. R$', 'MC R$', 'Resultado Real %', '% do Vendedor']
    ws.cell(row=title_row, column=RIGHT_COL0, value='GERENCIAL POR CLIENTE').font = _font(bold=True, size=12, color=HEADER_BG)
    for j, h in enumerate(RIGHT_HEADERS, start=RIGHT_COL0):
        _header_cell(ws, header_row, j, h)

    # Agrupa por cliente (ordenado por faturamento decrescente)
    cliente_data = {}
    clientes_ordem = []
    for it in vendor_itens:
        k = (it['cliente_codigo'], it['cliente_nome'])
        if k not in cliente_data:
            clientes_ordem.append(k)
            cliente_data[k] = {'fat': 0.0, 'custo': 0.0, 'qtd': 0.0}
        cliente_data[k]['fat']   += it['faturamento']
        cliente_data[k]['custo'] += it['custo_total']
        cliente_data[k]['qtd']   += it['qtd']
    clientes_ordem.sort(key=lambda k: -cliente_data[k]['fat'])

    for i, (cod, nome) in enumerate(clientes_ordem):
        r = first_data_row + i
        d = cliente_data[(cod, nome)]
        mc_rs_c, mc_pct_c, res_real_c = _calc(d['fat'], d['custo'])
        fat_unit_c   = d['fat']   / d['qtd'] if d['qtd'] else 0.0
        custo_unit_c = d['custo'] / d['qtd'] if d['qtd'] else 0.0
        pv_pct = d['fat'] / total_fat * 100 if total_fat else 0.0
        _text    (ws, r, RIGHT_COL0,     nome)
        _qty     (ws, r, RIGHT_COL0 + 1, d['qtd'])
        _money   (ws, r, RIGHT_COL0 + 2, round(d['fat'], 2))
        _money   (ws, r, RIGHT_COL0 + 3, round(fat_unit_c,   2))
        _money   (ws, r, RIGHT_COL0 + 4, round(custo_unit_c, 2))
        _money   (ws, r, RIGHT_COL0 + 5, round(mc_rs_c,  2))
        _pct_cell(ws, r, RIGHT_COL0 + 6, round(res_real_c,  2), res_real_c)
        pv = ws.cell(row=r, column=RIGHT_COL0 + 7, value=round(pv_pct, 2))
        pv.number_format = PCT_FMT; pv.border = BORDER; pv.font = _font()

    right_total_row = first_data_row + len(clientes_ordem)
    _text    (ws, right_total_row, RIGHT_COL0, 'TOTAL', bold=True)
    ws.cell(row=right_total_row, column=RIGHT_COL0).fill = PatternFill('solid', fgColor='E9F5EF')
    _qty     (ws, right_total_row, RIGHT_COL0 + 1, total_qtd)
    _money   (ws, right_total_row, RIGHT_COL0 + 2, round(total_fat, 2))
    _money   (ws, right_total_row, RIGHT_COL0 + 3, round(avg_fat_unit,   2))
    _money   (ws, right_total_row, RIGHT_COL0 + 4, round(avg_custo_unit, 2))
    _money   (ws, right_total_row, RIGHT_COL0 + 5, round(total_mc_rs,  2))
    _pct_cell(ws, right_total_row, RIGHT_COL0 + 6, round(total_res_real, 2), total_res_real)
    pv100 = ws.cell(row=right_total_row, column=RIGHT_COL0 + 7, value=100.0)
    pv100.number_format = PCT_FMT; pv100.border = BORDER; pv100.font = _font(bold=True)

    # Larguras
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 32
    for col in range(3, 9):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.column_dimensions[get_column_letter(RIGHT_COL0)].width = 28
    for col in range(RIGHT_COL0 + 1, RIGHT_COL0 + 8):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate
    return ws


# ---------------------------------------------------------------------------
# Entrada principal
# ---------------------------------------------------------------------------

def gerar_xlsx(parsed, output_path):
    itens        = parsed['itens']
    data_emissao = parsed.get('data_emissao')
    periodo      = parsed.get('periodo')

    # Agrupa por vendedor
    fat_por_vendedor   = {}
    mc_por_vendedor    = {}
    res_por_vendedor   = {}
    itens_por_vendedor = {}
    for it in itens:
        vname = it['vendedor'] or it['vendedor_raw']
        itens_por_vendedor.setdefault(vname, []).append(it)
        fat_por_vendedor[vname] = fat_por_vendedor.get(vname, 0.0) + it['faturamento']

    for vname, rows in itens_por_vendedor.items():
        fat   = sum(r['faturamento'] for r in rows)
        custo = sum(r['custo_total'] for r in rows)
        _, mc_pct, res_real = _calc(fat, custo)
        mc_por_vendedor[vname]  = mc_pct
        res_por_vendedor[vname] = res_real

    vendor_order = sorted(fat_por_vendedor.keys(), key=lambda v: -fat_por_vendedor[v])
    vendor_order_with_pct = [(v, mc_por_vendedor[v], res_por_vendedor[v]) for v in vendor_order]

    wb = Workbook()
    wb.remove(wb.active)

    _build_leiame(wb, data_emissao, periodo)
    _, n_dados = _build_dados(wb, itens)
    _build_resumo(wb, n_dados, vendor_order_with_pct)
    _build_alertas(wb, itens)
    for vname in vendor_order:
        _build_vendedor_sheet(wb, vname, itens_por_vendedor[vname])

    wb.active = 1
    wb.save(output_path)
    return output_path
