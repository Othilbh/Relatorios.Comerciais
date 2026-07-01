"""Gerador do Excel de Recorrencia de Vendas OTHIL.

Fonte: PDF 'Lucratividade por Vendedor-Cliente no Previsao' (Mercatus).
Reutiliza parsers_diario.parse_relatorio_diario para extracao dos dados.

Estrutura do Excel:
  - Aba CONSOLIDADO (todos os vendedores exceto Luca)
  - Uma aba por vendedor ativo (exceto Luca)
  Cada aba: matriz cliente x produto-unificado em caixas (CX)
  Colunas fixas: CLIENTE | TOTAL CX | FATURAMENTO (R$) | MARGEM REL. | MARGEM REAL +15%
  Colunas de produto ordenadas por volume total (CX) desc
  Verde  = cliente comprou esse produto no periodo
  Laranja = produto no mix mas NAO comprado pelo cliente
"""
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

DESP_ADM = 1.15

C_HEADER  = '1B2A4A'
C_VERDE_H = '2D6A4F'
C_VERDE   = 'D8EFE3'
C_LARANJA = 'FFE5CC'
C_LAR_FG  = 'C75B00'
C_NEG     = 'C0392B'

EXCLUIR = {'Luca'}

# =============================================================================
# Consolidacao de clientes
# =============================================================================

_SUFIXOS_WPP = [
    ' / WHATSAPP', '/ WHATSAPP', ' - WHATSAPP', ' WHATSAPP',
    '/WHATSAPP', ' / WPP', '/ WPP', ' - WPP', ' WPP',
    ' / ZAP', '/ ZAP', ' - ZAP',
]

# Grupos explicitamente definidos (chave = substring a buscar no nome UPPER)
_GRUPOS = [
    ('MERCANTIL BASTOS',   'MERCANTIL BASTOS'),
    ('MERC BASTOS',        'MERCANTIL BASTOS'),
    ('ENTRE RIOS',         'ENTRE RIOS'),
    ('PAIS E FILHOS',      'PAIS E FILHOS'),
    ('PAI E FILHOS',       'PAIS E FILHOS'),
    ('NOVA SU',            'DARIO'),
    ('SANTA LU',           'DARIO'),
    ('SINFR',              'DARIO'),
    ('ADELSON BOY',        'ADELSON BOY'),
    ('TEREZINHA',          'TEREZINHA'),
    ('GOMES & SANTOS',     'GOMES & SANTOS'),
    ('NOVO HORIZONTE',     'NOVO HORIZONTE'),
    ('SACOLAO CASTELO',    'SACOLAO CASTELO'),
    ('SACOLAO MAXIMO',     'SACOLAO MAXIMO'),
    ('ABC CAMPEAO',        'ABC CAMPEAO'),
    ('EXPERFRUT',          'EXPERFRUT'),
    ('MACHACALIS',         'ANDRE MACHACALIS'),
    # Grupos com sufixo operacional (detectados antes de qualquer stripping)
    ('VILLEFORT',          'VILLEFORT'),
    ('BIG MAIS',           'BIG MAIS'),
    ('FRUTAS BARBACENA',   'FRUTAS BARBACENA'),
    ('MART MINAS',         'MART MINAS'),
    ('BRIXX',              'BRIXX'),
    ('BONFIM',             'BONFIM'),
]

_RE_G = re.compile(r'GOMES')
_RE_S = re.compile(r'SANTOS')


def consolidar_cliente(nome: str) -> str:
    n = nome.upper().strip()

    # 1. Remove sufixos WhatsApp
    for sfx in _SUFIXOS_WPP:
        if sfx in n:
            n = n.replace(sfx, '').strip()

    # 2. Remove "- CD" (centro de distribuicao)
    n = re.sub(r'\s*-\s*CD\s*$', '', n).strip()

    # 3. Remove "- PAV X" ou " PAV X" (pavilhao, com ou sem traco)
    n = re.sub(r'\s*-?\s*PAV\b.*$', '', n).strip()

    # 4. Remove "- <cidade>" no final (ex: "- JUIZ DE FORA")
    n = re.sub(r'\s+-\s+[A-Z\s]{3,}$', '', n).strip()

    # 5. Regra especial: GOMES + SANTOS (qualquer ordem)
    if _RE_G.search(n) and _RE_S.search(n):
        return 'GOMES & SANTOS'

    # 6. Grupos explicitos
    for chave, canonico in _GRUPOS:
        if chave in n:
            return canonico

    return n


# =============================================================================
# Mapeamento produto -> coluna unificada
# =============================================================================

def _col(descricao: str) -> str:
    d = descricao.upper()

    # Macas
    if re.search(r'\bGALA\b', d):
        return 'MACA GALA'
    if re.search(r'\bFUJI\b', d):
        return 'MACA FUJI'
    if re.search(r'\b(ARGENTINA|CHILENA|PINK LADY|GRAN.?SMITH|GRANNY.?SMITH)\b', d):
        return 'MACA ARGENTINA'

    # Peras — captura WILLIAMS, WILLIAM, WILLIANS, WILLAMS, WILL (abrev.), PACKHAM
    if re.search(r'\bWILL(IANS?|AMS?|IAM)?S?\b|\bPACKHAMS?\b', d) and 'BANANA' not in d:
        return 'PERA WILLIAMS'
    if re.search(r'\bPORTUGUESA\b', d) and 'UVA' not in d:
        return 'PERA PORTUGUESA'
    if re.search(r'\b(FORELLE|ERCOLINE|ASIATICA|ASIATICA)\b', d):
        return 'FORELLE / ERCOLINE'

    # Uvas — especificas antes da generica
    if re.search(r'\bTHOMPSON\b', d):
        return 'UVA THOMPSON'
    if re.search(r'\bVITORIA\b', d):
        return 'UVA VITORIA'
    if re.search(r'\b(ISIS|CRIN?SON|NUBIA|JUBIL[EL]{1,2})\b', d):
        return 'UVA VERMELHA'
    if re.search(r'\b(ROSADA|RANDA)\b', d):
        return 'UVA ROSADA / RANDA'
    if re.search(r'\b(BENITAKA|RED GLOBE|ITALIA)\b', d):
        return 'UVA 8 KG'
    if re.search(r'\bCUMB\b', d):
        return 'UVA'

    # Kiwi
    if re.search(r'\bKIWI\b', d):
        return 'KIWI'

    # Melancia / Melao
    if re.search(r'\bMELANCIA\b', d):
        return 'MELANCIA'
    if re.search(r'\bMELAO\b', d):
        return 'MELAO'

    # Caqui
    if re.search(r'\bCAQUI\b', d):
        return 'CAQUI'

    # Caroco (ameixa + pessego)
    if re.search(r'\b(AMEIXA|PESSEGO|NECTARINA)\b', d):
        return 'CAROCO'

    # Mamao
    if re.search(r'\b(MAMAO|HAVAI|FORMOSO)\b', d):
        return 'MAMAO'

    # Citros
    if re.search(r'\b(MEXERICA|TANGERINA|BERGAMOTA)\b', d):
        return 'MEXERICA / TANGERINA'
    if re.search(r'\bLIMAO\b', d):
        return 'LIMAO'
    if re.search(r'\bLARANJA\b', d):
        return 'LARANJA'

    # Roma / Mirtilo
    if re.search(r'\b(ROMA|MIRTILO)\b', d):
        return 'ROMA / MIRTILO'

    # Outros
    if re.search(r'\bGOIABA\b', d):
        return 'GOIABA'
    if re.search(r'\bTAMARA\b', d):
        return 'TAMARA'
    if re.search(r'\bMARACUJA\b', d):
        return 'MARACUJA'
    if re.search(r'\bPIMENTAO?\b|\bPIMENTA\b', d):
        return 'PIMENTA'
    if re.search(r'\bLASANHA\b', d):
        return 'LASANHA'

    # Fallback: descricao como coluna propria
    return d.strip()


# Nomes de exibicao para colunas conhecidas
_DISPLAY = {
    'MACA GALA':            'MACA GALA',
    'MACA FUJI':            'MACA FUJI',
    'MACA ARGENTINA':       'MACA ARGENTINA',
    'PERA WILLIAMS':        'PERA WILLIAMS',
    'PERA PORTUGUESA':      'PERA PORTUGUESA',
    'FORELLE / ERCOLINE':   'FORELLE / ERCOLINE',
    'UVA THOMPSON':         'UVA THOMPSON',
    'UVA VITORIA':          'UVA VITORIA',
    'UVA VERMELHA':         'UVA VERMELHA',
    'UVA ROSADA / RANDA':   'UVA ROSADA / RANDA',
    'UVA 8 KG':             'UVA 8 KG',
    'UVA':                  'UVA',
    'KIWI':                 'KIWI',
    'MELANCIA':             'MELANCIA',
    'MELAO':                'MELAO',
    'CAQUI':                'CAQUI',
    'CAROCO':               'CAROCO',
    'MAMAO':                'MAMAO',
    'MEXERICA / TANGERINA': 'MEXERICA / TANGERINA',
    'LIMAO':                'LIMAO',
    'LARANJA':              'LARANJA',
    'ROMA / MIRTILO':       'ROMA / MIRTILO',
    'GOIABA':               'GOIABA',
    'TAMARA':               'TAMARA',
    'MARACUJA':             'MARACUJA',
    'PIMENTA':              'PIMENTA',
    'LASANHA':              'LASANHA',
}


# =============================================================================
# Agregacao
# =============================================================================

def _agregar(itens: list) -> dict:
    """Retorna {cliente_consolidado: {'_fat', '_custo', col: qty, ...}}"""
    out = {}
    for it in itens:
        cli = consolidar_cliente(it['cliente_nome'])
        d = out.setdefault(cli, {'_fat': 0.0, '_custo': 0.0})
        d['_fat']   += it['faturamento']
        d['_custo'] += it['custo_total']
        col = _col(it['produto'])
        d[col] = d.get(col, 0.0) + it['qtd']
    return out


def _colunas_ordenadas(clientes_dict: dict) -> list:
    """Colunas de produto ordenadas por CX total desc (apenas > 0)."""
    totais: dict = {}
    for d in clientes_dict.values():
        for k, v in d.items():
            if not k.startswith('_'):
                totais[k] = totais.get(k, 0.0) + v
    return [k for k, v in sorted(totais.items(), key=lambda x: -x[1]) if v > 0]


# =============================================================================
# Periodo
# =============================================================================

_MESES = ['', 'JANEIRO', 'FEVEREIRO', 'MARCO', 'ABRIL', 'MAIO', 'JUNHO',
          'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']


def _titulo_periodo(s: str) -> str:
    if not s:
        return ''
    datas = re.findall(r'(\d{2})/(\d{2})/(\d{4})', s)
    if len(datas) < 2:
        return s
    d1, m1, y1 = datas[0]
    d2, m2, y2 = datas[1]
    m = _MESES[int(m1)] if m1 == m2 else _MESES[int(m1)] + '/' + _MESES[int(m2)]
    if d1 == '01' and int(d2) >= 28:
        label = m + '/' + y1
    elif int(d2) <= 15:
        label = '1a quinzena ' + m + '/' + y1
    else:
        label = '2a quinzena ' + m + '/' + y1
    return label + ' (' + d1 + '/' + m1 + ' a ' + d2 + '/' + m2 + '/' + y2 + ')'


def _filename_periodo(s: str) -> str:
    if not s:
        return 'periodo'
    datas = re.findall(r'(\d{2})/(\d{2})/(\d{4})', s)
    if len(datas) >= 2:
        d1, m1, y1 = datas[0]
        d2, m2, y2 = datas[1]
        if y1 == y2 and m1 == m2:
            return d1 + 'a' + d2 + '-' + m1 + '-' + y1
        return d1 + '-' + m1 + '-' + y1 + '_a_' + d2 + '-' + m2 + '-' + y2
    return re.sub(r'[^A-Za-z0-9_-]', '_', s)[:30]


# =============================================================================
# Excel writer
# =============================================================================

_HDR   = ['CLIENTE', 'TOTAL CX', 'FATURAMENTO (R$)', 'MARGEM REL.', 'MARGEM REAL +15%']
_N_HDR = len(_HDR)


def _preencher_aba(ws, clientes_dict: dict, prod_cols: list,
                   titulo: str, periodo_str: str):

    n_cols = _N_HDR + len(prod_cols)

    fill_nav = PatternFill('solid', fgColor=C_HEADER)
    fill_vh  = PatternFill('solid', fgColor=C_VERDE_H)
    fill_vc  = PatternFill('solid', fgColor=C_VERDE)
    fill_lar = PatternFill('solid', fgColor=C_LARANJA)

    font_title  = Font(name='Arial', color='FFFFFF', bold=True, size=14)
    font_legend = Font(name='Arial', size=8, italic=True, color='444444')
    font_hdr    = Font(name='Arial', color='FFFFFF', bold=True, size=9)
    font_data   = Font(name='Arial', size=9)
    font_neg    = Font(name='Arial', size=9, color=C_NEG, bold=True)
    font_lar    = Font(name='Arial', size=9, color=C_LAR_FG)
    font_tot    = Font(name='Arial', color='FFFFFF', bold=True, size=9)
    font_cli    = Font(name='Arial', size=9, bold=True)

    al_cen = Alignment(horizontal='center', vertical='center')
    al_lft = Alignment(horizontal='left',   vertical='center')
    al_rgt = Alignment(horizontal='right',  vertical='center')
    al_rot = Alignment(text_rotation=90, horizontal='center', vertical='bottom')
    al_wlf = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Linha 1 — Titulo
    ws.row_dimensions[1].height = 42
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1, 'OTHIL - RECORRENCIA DE VENDAS  |  ' + titulo)
    c.fill, c.font, c.alignment = fill_nav, font_title, al_cen

    # Linha 2 — Legenda
    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    leg = (
        'Verde = vendido  |  Laranja = NAO vendido (disponivel no mix OTHIL)  |  '
        'Valores = caixas  |  Margem Real = Margem Relatorio + 15% (custo operacional)'
        '  |  Periodo: ' + periodo_str
    )
    c = ws.cell(2, 1, leg)
    c.font, c.alignment = font_legend, al_wlf

    # Linha 3 — Cabecalhos de coluna
    ws.row_dimensions[3].height = 95
    for i, h in enumerate(_HDR, 1):
        c = ws.cell(3, i, h)
        c.fill, c.font = fill_vh, font_hdr
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for j, col in enumerate(prod_cols, _N_HDR + 1):
        c = ws.cell(3, j, _DISPLAY.get(col, col))
        c.fill, c.font, c.alignment = fill_vh, font_hdr, al_rot

    def _tcx(d):
        return sum(v for k, v in d.items() if not k.startswith('_'))

    rows_sorted = sorted(clientes_dict.items(), key=lambda x: -_tcx(x[1]))

    # Linhas de dados
    for row_n, (cli, d) in enumerate(rows_sorted, start=4):
        ws.row_dimensions[row_n].height = 15
        fat   = d['_fat']
        custo = d['_custo']
        cr    = custo / DESP_ADM if custo else 0.0
        mc    = fat - cr
        m_rel  = (fat - custo) / custo * 100 if custo else 0.0
        m_real = mc / cr * 100 if cr else 0.0
        tcx   = _tcx(d)

        c = ws.cell(row_n, 1, cli)
        c.font, c.alignment = font_cli, al_lft

        c = ws.cell(row_n, 2, round(tcx))
        c.number_format = '#,##0'
        c.fill, c.font, c.alignment = fill_vc, font_data, al_rgt

        c = ws.cell(row_n, 3, round(fat, 2))
        c.number_format = '#,##0.00'
        c.font, c.alignment = font_data, al_rgt

        c = ws.cell(row_n, 4, round(m_rel / 100, 4))
        c.number_format = '0.00%'
        c.font = font_neg if m_rel < 0 else font_data
        c.alignment = al_rgt

        c = ws.cell(row_n, 5, round(m_real / 100, 4))
        c.number_format = '0.00%'
        c.font = font_neg if m_real < 0 else font_data
        c.alignment = al_rgt

        for j, col in enumerate(prod_cols, _N_HDR + 1):
            qty = d.get(col, 0.0)
            if qty > 0:
                c = ws.cell(row_n, j, round(qty))
                c.number_format = '#,##0'
                c.fill, c.font, c.alignment = fill_vc, font_data, al_rgt
            else:
                c = ws.cell(row_n, j, '-')
                c.fill, c.font, c.alignment = fill_lar, font_lar, al_cen

    # Linha TOTAL GERAL
    tot_row = 4 + len(rows_sorted)
    ws.row_dimensions[tot_row].height = 18

    fat_t   = sum(d['_fat']   for _, d in rows_sorted)
    cst_t   = sum(d['_custo'] for _, d in rows_sorted)
    cr_t    = cst_t / DESP_ADM if cst_t else 0.0
    mc_t    = fat_t - cr_t
    mr_t    = (fat_t - cst_t) / cst_t * 100 if cst_t else 0.0
    mreal_t = mc_t / cr_t * 100 if cr_t else 0.0
    tcx_t   = sum(_tcx(d) for _, d in rows_sorted)

    c = ws.cell(tot_row, 1, 'TOTAL GERAL')
    c.fill, c.font, c.alignment = fill_vh, font_tot, al_lft

    c = ws.cell(tot_row, 2, round(tcx_t))
    c.number_format = '#,##0'
    c.fill, c.font, c.alignment = fill_vh, font_tot, al_rgt

    c = ws.cell(tot_row, 3, round(fat_t, 2))
    c.number_format = '#,##0.00'
    c.fill, c.font, c.alignment = fill_vh, font_tot, al_rgt

    c = ws.cell(tot_row, 4, round(mr_t / 100, 4))
    c.number_format = '0.00%'
    c.fill, c.font, c.alignment = fill_vh, font_tot, al_rgt

    c = ws.cell(tot_row, 5, round(mreal_t / 100, 4))
    c.number_format = '0.00%'
    c.fill, c.font, c.alignment = fill_vh, font_tot, al_rgt

    for j, col in enumerate(prod_cols, _N_HDR + 1):
        qty = sum(d.get(col, 0.0) for _, d in rows_sorted)
        c = ws.cell(tot_row, j, round(qty) if qty > 0 else '')
        if qty > 0:
            c.number_format = '#,##0'
        c.fill, c.font, c.alignment = fill_vh, font_tot, al_rgt

    # Larguras
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    for j in range(_N_HDR + 1, n_cols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 7.5

    ws.freeze_panes = ws.cell(4, _N_HDR + 1)


# =============================================================================
# Entry point
# =============================================================================

def gerar_xlsx(parsed: dict, output_path: str) -> str:
    """Gera o Excel e salva em output_path. Retorna string de periodo."""
    itens   = parsed['itens']
    periodo = parsed.get('periodo', '')
    titulo  = _titulo_periodo(periodo)

    itens_ok = [it for it in itens if it.get('vendedor') not in EXCLUIR]

    cons_dict = _agregar(itens_ok)
    prod_cols = _colunas_ordenadas(cons_dict)

    wb = Workbook()

    ws_cons = wb.active
    ws_cons.title = 'CONSOLIDADO'
    _preencher_aba(ws_cons, cons_dict, prod_cols, titulo, periodo)

    por_vend: dict = {}
    for it in itens_ok:
        v = it.get('vendedor') or it.get('vendedor_raw', '?')
        por_vend.setdefault(v, []).append(it)

    for vname in sorted(por_vend.keys()):
        vd = _agregar(por_vend[vname])
        ws = wb.create_sheet(title=vname[:31])
        _preencher_aba(ws, vd, prod_cols, titulo, periodo)

    wb.save(output_path)
    return _filename_periodo(periodo)
